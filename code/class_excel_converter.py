import datetime

from class_time_line import TimeLine
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

from class_location import Location
import openpyxl


class ExcelConverter:
    DEFAULT_SAVE_PATH = "../csv_data/trip_plan_gwd.xlsx"
    # 가운데 정렬
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')

    # 글씨체 굵게
    FONT_BOLD = Font(size=14, bold=True, color='000000')  # 000000: black
    FONT_TITLE = Font(size=16, bold=True, color='FFFFFF')
    FONT_SUB_TITLE = Font(size=14, bold=True, color='000000')
    FONT_PLAIN = Font(size=14, color='000000')

    # 셀 색깔 채우기
    FILL_LIME = PatternFill('solid', fgColor='92F56B')
    FILL_BLUE = PatternFill('solid', fgColor='1BA8FB')
    FILL_BROWN = PatternFill('solid', fgColor='3B240B')
    FILL_GREEN = PatternFill('solid', fgColor='A9F5A9')
    FILL_IVORY = PatternFill('solid', fgColor='CECEF6')
    FILL_LIGHT_YELLOW = PatternFill('solid', fgColor='F7F8E0')

    # 테두리 선넣기
    THIN_BORDER = Border(left=Side(border_style='thin', color='000000'),
                         right=Side(border_style='thin', color='000000'),
                         top=Side(border_style='thin', color='000000'),
                         bottom=Side(border_style='thin', color='000000'))

    def __init__(self, timeline_obj: TimeLine = None):
        self.time_line = timeline_obj
        self.work_book = None
        self.worksheet = None
        self.create_workbook()

    def create_workbook(self):
        # 워크북 생성
        self.work_book = openpyxl.Workbook()
        wb = self.work_book
        # 시트 활성화
        self.worksheet = wb.worksheets[0]
        ws = self.worksheet
        # 시트 이름바꾸기
        ws.title = '여행계획'
        # 셀 높이 조정
        for row in range(1, 20):
            ws.row_dimensions[row].height = 30
        # 셀 넓이 조정
        col_width_list = [10, 10, 24, 35, 50]
        col_idx = 0
        for col in range(65, 70):
            ws.column_dimensions[chr(col)].width = col_width_list[col_idx]
            col_idx += 1

        # 전체 가운데 정렬
        for row in ws['A1':'E30']:
            for cell in row:
                cell.alignment = self.ALIGN_CENTER

        # 여행 헤드 타이틀 서식 설정
        ws['B1'].value = '일정'
        ws['D1'].value = '"봄 감자가 맛있단다"'

        ws['B1'].font = self.FONT_BOLD

        for row in ws['B1':'C1']:
            for cell in row:
                cell.border = self.THIN_BORDER

        for row in ws['D1':'E1']:
            for cell in row:
                cell.font = self.FONT_BOLD
        ws['E1'].fill = self.FILL_BROWN

    def set_day_count_format(self, cell):
        cell.fill = self.FILL_BLUE
        cell.font = self.FONT_SUB_TITLE
        cell.border = self.THIN_BORDER

    def set_sub_category_format(self, cell):
        cell.font = self.FONT_BOLD
        cell.fill = self.FILL_LIME
        cell.border = self.THIN_BORDER

    def set_fill_plain(self, cell):
        cell.font = self.FONT_PLAIN
        cell.border = self.THIN_BORDER

    def add_row_day_and_category_title(self, row_index, offset, time_line_obj):
        # 변수, 함수 설정
        self.time_line = time_line_obj
        time_line = time_line_obj
        start_date = time_line.plan_date.start_date
        date_obj_to_str = time_line.plan_date.date_obj_to_str
        current_day = start_date + datetime.timedelta(days=offset)

        ws = self.worksheet

        # 데이 카운터 서식 설정
        ws.cell(row=row_index, column=2).value = f'Day{offset + 1}'
        ws.cell(row=row_index, column=3).value = date_obj_to_str(current_day)
        self.set_day_count_format(ws.cell(row=row_index, column=2))
        self.set_day_count_format(ws.cell(row=row_index, column=3))

        # 분류 타이틀 문자 설정
        ws[f'B{row_index + 1}'].value = '유형'
        ws[f'C{row_index + 1}'].value = '장소'
        ws[f'D{row_index + 1}'].value = '주소'
        ws[f'E{row_index + 1}'].value = '메모'

        # 분류 타이틀 서식 설정
        self.set_sub_category_format(ws[f'B{row_index + 1}'])
        self.set_sub_category_format(ws[f'C{row_index + 1}'])
        self.set_sub_category_format(ws[f'D{row_index + 1}'])
        self.set_sub_category_format(ws[f'E{row_index + 1}'])

    def save_excel_file(self, path=None):
        wb = self.work_book
        if path is not None:
            wb.save(f"{path}.xlsx")
        else:
            wb.save(self.DEFAULT_SAVE_PATH)

    def set_timeline_duration_str(self):
        ws = self.worksheet
        date_obj_to_str = self.time_line.plan_date.date_obj_to_str
        start_date = self.time_line.plan_date.start_date
        end_date = self.time_line.plan_date.end_date
        start_date_str = date_obj_to_str(start_date)
        end_date_str = date_obj_to_str(end_date)
        ws['C1'].value = f"{start_date_str} ~ {end_date_str}"

    def set_timeline(self, time_line_obj: TimeLine):
        print(time_line_obj)
        # 변수, 함수 설정
        self.time_line = time_line_obj
        time_line = self.time_line
        current_row_idx = 2

        ws = self.worksheet
        ws['E1'].font = self.FONT_TITLE
        ws['E1'].value = time_line.trip_name

        self.set_timeline_duration_str()

        for offset, daily_location_list in enumerate(time_line.location_list):
            self.add_row_day_and_category_title(current_row_idx, offset, time_line_obj)
            current_row_idx += 2
            if len(daily_location_list) == 0:
                self.add_location_row(current_row_idx, 0, None)
                current_row_idx += 2
                continue

            for inner_idx, location_obj in enumerate(daily_location_list):
                inner_idx += 1
                self.add_location_row(current_row_idx, inner_idx, location_obj)
                current_row_idx += 1
            current_row_idx += 1

    def add_location_row(self, row_idx, inner_idx, location_obj: Location = None):
        ws = self.worksheet
        if location_obj is not None:
            category = location_obj.category
            name = location_obj.name
            address = location_obj.address

        else:
            category = ''
            name = ''
            address = ''

        if category == '명소' or category == 1:
            ws[f'A{row_idx}'].value = inner_idx
            ws[f'A{row_idx}'].fill = self.FILL_IVORY
            ws[f'B{row_idx}'].value = '명소'

        elif category == '숙소'  or category == 0:
            ws[f'A{row_idx}'].value = '*'
            ws[f'A{row_idx}'].fill = self.FILL_GREEN
            ws[f'B{row_idx}'].value = '숙박'

        else:
            ws[f'A{row_idx}'].fill = self.FILL_LIGHT_YELLOW

        ws[f'C{row_idx}'].value = f'{name}'
        ws[f'D{row_idx}'].value = f'{address}'

        self.set_fill_plain(ws[f'B{row_idx}'])
        self.set_fill_plain(ws[f'C{row_idx}'])
        self.set_fill_plain(ws[f'D{row_idx}'])
        self.set_fill_plain(ws[f'E{row_idx}'])
